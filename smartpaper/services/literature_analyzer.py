"""
文獻分析服務
提供三大功能：
1. 自動大綱生成 (generate_outline) — 主題 → 段落結構 + 每段推薦引用
2. 文獻回顧表格 (generate_review_table) — LLM 從 abstract 萃取欄位 → Excel
3. 論文差異比較 (compare_papers) — 多篇論文跨維度比較表
"""

import json
from dataclasses import dataclass, field
from typing import Optional, Callable
from google import genai

from ..database.sqlite_db import SQLiteDB
from ..database.vector_db import VectorDB
from ..config import GEMINI_API_KEY, GEMINI_MODEL
from ..models import Paper
from .reranker import Reranker


# ── Outline generation ────────────────────────────────────────────────────

@dataclass
class OutlinePaper:
    paper: Paper
    cite_reason: str
    cite_position: str   # 開頭 / 中間 / 結尾
    key_concept: str


@dataclass
class OutlineSection:
    title: str            # 段落標題，如「Introduction」
    description: str      # 段落說明
    writing_hint: str     # 寫作提示
    papers: list[OutlinePaper] = field(default_factory=list)


# ── Review table ──────────────────────────────────────────────────────────

# 固定欄位（從 DB 直接取得，不需 LLM）
FIXED_COLUMNS = {
    "標題":   lambda p: p.title,
    "作者":   lambda p: "; ".join(p.authors[:3]) + (" et al." if len(p.authors) > 3 else "") if p.authors else "-",
    "年份":   lambda p: str(p.year) if p.year else "-",
    "期刊/會議": lambda p: p.venue or "-",
    "DOI":   lambda p: p.doi or "-",
    "引用數": lambda p: str(p.citation_count) if p.citation_count else "-",
    "標籤":  lambda p: ", ".join(p.tags[:4]) if p.tags else "-",
}

# LLM 可萃取的欄位（預設問法）
LLM_COLUMNS = {
    "研究問題":  "這篇論文試圖解決什麼研究問題？（20字以內）",
    "研究方法":  "這篇論文使用了什麼方法或模型？（20字以內）",
    "資料集":   "論文使用了哪些資料集？若未提及請回答「未說明」（15字以內）",
    "主要貢獻":  "論文最主要的貢獻或發現是什麼？（25字以內）",
    "主要結果":  "論文的主要實驗結果或效果如何？（20字以內）",
    "研究限制":  "論文的局限性或未來工作方向是什麼？（20字以內）",
    "適用場景":  "這個研究成果適合應用在哪些場景？（20字以內）",
}


class LiteratureAnalyzer:
    """文獻分析服務"""

    def __init__(
        self,
        sqlite_db: Optional[SQLiteDB] = None,
        vector_db: Optional[VectorDB] = None,
        api_key: Optional[str] = None,
    ):
        self.db = sqlite_db or SQLiteDB()
        self.vector_db = vector_db or VectorDB()
        self.reranker = Reranker()
        self.api_key = api_key or GEMINI_API_KEY
        if self.api_key:
            self.client = genai.Client(api_key=self.api_key)
        else:
            self.client = None

    # ── 1. 自動大綱生成 ──────────────────────────────────────────────────

    def generate_outline(
        self,
        topic: str,
        n_candidates: int = 15,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> list[OutlineSection]:
        """
        根據研究主題自動生成論文大綱，並為每個段落推薦引用。

        流程：
        1. LLM 根據主題建議大綱結構（段落清單）
        2. 對每個段落做語意搜尋 + re-rank 找候選論文
        3. LLM 分析候選論文，為每段決定應引用哪些及原因

        Returns:
            list[OutlineSection]
        """
        if not self.client:
            raise ValueError("需要 Gemini API Key")

        if progress_callback:
            progress_callback("生成大綱結構...")

        # Step 1: 讓 LLM 建議段落結構
        sections_raw = self._suggest_sections(topic)

        results: list[OutlineSection] = []

        for i, sec in enumerate(sections_raw):
            if progress_callback:
                progress_callback(f"分析段落 [{i+1}/{len(sections_raw)}] {sec['title']}...")

            # Step 2: 語意搜尋 + re-rank
            query = f"{sec['title']} {sec['description']} {topic}"
            candidates = self._retrieve_candidates(query, n_candidates)

            # Step 3: LLM 決定引用
            outline_papers = self._assign_citations(topic, sec, candidates)

            results.append(OutlineSection(
                title=sec["title"],
                description=sec["description"],
                writing_hint=sec.get("writing_hint", ""),
                papers=outline_papers,
            ))

        return results

    def _suggest_sections(self, topic: str) -> list[dict]:
        """讓 LLM 建議論文段落結構"""
        prompt = f"""你是一位學術寫作專家。請為主題「{topic}」建議一個學術論文的段落結構。

請生成 5-7 個段落，每個段落包含：
- 段落標題（英文，如 Introduction, Related Work, Methodology...）
- 段落說明（這個段落應該寫什麼）
- 寫作提示（關鍵要點）

以 JSON 格式回答：
{{
  "sections": [
    {{
      "title": "Introduction",
      "description": "介紹研究背景、動機與研究問題",
      "writing_hint": "從廣到窄，先介紹領域背景，再收斂到具體研究問題，最後說明本研究的貢獻"
    }},
    ...
  ]
}}

只回傳 JSON。"""

        resp = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
        text = resp.text.strip()
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]
        data = json.loads(text.strip())
        return data.get("sections", [])

    def _retrieve_candidates(self, query: str, n: int) -> list[Paper]:
        """語意搜尋 + re-rank 取得候選論文"""
        vec_results = self.vector_db.search(query=query, n_results=n * 2)
        papers = []
        seen = set()
        for vr in vec_results:
            p = self.db.get_by_id(vr["paper_id"])
            if p and p.id not in seen:
                seen.add(p.id)
                papers.append(p)

        if not papers:
            return []

        try:
            rerank_inputs = [
                {"paper": p, "document": f"{p.title}. {(p.abstract or '')[:300]}"}
                for p in papers
            ]
            reranked = self.reranker.rerank(query, rerank_inputs, text_key="document", top_k=n)
            return [r["paper"] for r in reranked]
        except Exception:
            return papers[:n]

    def _assign_citations(self, topic: str, section: dict, candidates: list[Paper]) -> list[OutlinePaper]:
        """LLM 決定哪些論文應在此段落引用"""
        if not candidates:
            return []

        papers_ctx = "\n".join(
            f"[{i+1}] 標題：{p.title}\n    摘要：{(p.abstract or '無摘要')[:300]}"
            for i, p in enumerate(candidates)
        )

        prompt = f"""研究主題：{topic}
當前段落：{section['title']} — {section['description']}

以下是候選論文（編號對應論文）：
{papers_ctx}

請判斷哪些論文適合在「{section['title']}」段落中引用。

對每篇應引用的論文，請提供：
- 引用原因（20字以內）
- 引用位置（開頭 / 中間 / 結尾）
- 核心概念（從論文中提取，15字以內）

JSON 格式：
{{
  "citations": [
    {{
      "paper_index": 1,
      "cite_reason": "提供研究背景",
      "cite_position": "開頭",
      "key_concept": "注意力機制的定義"
    }}
  ]
}}

只引用真正相關的論文（可以是 0 篇），只回傳 JSON。"""

        try:
            resp = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            text = resp.text.strip()
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]
            data = json.loads(text.strip())
            citations = data.get("citations", [])

            result = []
            for c in citations:
                idx = c.get("paper_index", 0) - 1
                if 0 <= idx < len(candidates):
                    result.append(OutlinePaper(
                        paper=candidates[idx],
                        cite_reason=c.get("cite_reason", ""),
                        cite_position=c.get("cite_position", "中間"),
                        key_concept=c.get("key_concept", ""),
                    ))
            return result
        except Exception as e:
            print(f"引用分配失敗: {e}")
            return []

    # ── 2. 文獻回顧表格 ──────────────────────────────────────────────────

    def generate_review_table(
        self,
        papers: list[Paper],
        fixed_cols: list[str],
        llm_cols: list[str],
        custom_cols: Optional[list[str]] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> list[dict]:
        """
        生成文獻回顧表格。

        Args:
            papers: 要分析的論文
            fixed_cols: 從 DB 直接取的欄位（見 FIXED_COLUMNS）
            llm_cols: 用 LLM 從 abstract 萃取的欄位（見 LLM_COLUMNS）
            custom_cols: 使用者自訂欄位名稱（LLM 嘗試回答）
            progress_callback: (current, total)

        Returns:
            [{"標題": ..., "作者": ..., "研究方法": ..., ...}, ...]
        """
        all_llm_cols = list(llm_cols) + (custom_cols or [])
        rows = []

        for i, paper in enumerate(papers):
            if progress_callback:
                progress_callback(i + 1, len(papers))

            row: dict = {}

            # 固定欄位（直接取）
            for col in fixed_cols:
                if col in FIXED_COLUMNS:
                    row[col] = FIXED_COLUMNS[col](paper)

            # LLM 欄位（批次萃取）
            if all_llm_cols and paper.abstract:
                extracted = self._extract_fields(paper, all_llm_cols, custom_cols or [])
                row.update(extracted)
            elif all_llm_cols:
                for col in all_llm_cols:
                    row[col] = "（無摘要）"

            rows.append(row)

        return rows

    def _extract_fields(
        self,
        paper: Paper,
        llm_cols: list[str],
        custom_cols: list[str],
    ) -> dict:
        """一次 LLM call 萃取所有欄位"""
        questions = []
        for col in llm_cols:
            if col in LLM_COLUMNS:
                questions.append(f'"{col}": {LLM_COLUMNS[col]}')
            else:
                questions.append(f'"{col}": 請根據摘要回答「{col}」，20字以內，若無法判斷請回答「未說明」')

        questions_str = "\n".join(f"- {q}" for q in questions)

        prompt = f"""請根據以下論文摘要，回答各個問題。

論文標題：{paper.title}
論文摘要：{(paper.abstract or '')[:800]}

需要回答的問題：
{questions_str}

以 JSON 格式回答（每個欄位值不超過規定字數）：
{{{", ".join(f'"{col}": "..."' for col in llm_cols)}}}

只回傳 JSON。"""

        try:
            resp = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            text = resp.text.strip()
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]
            return json.loads(text.strip())
        except Exception:
            return {col: "萃取失敗" for col in llm_cols}

    def save_review_table_xlsx(
        self,
        rows: list[dict],
        output_path: str,
        title: str = "文獻回顧",
    ) -> None:
        """將文獻回顧表格存成 Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        if not rows:
            wb.save(output_path)
            return

        headers = list(rows[0].keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, "-"))
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 自動欄寬
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 30
        wb.save(output_path)

    # ── 3. 論文差異比較 ──────────────────────────────────────────────────

    def compare_papers(
        self,
        papers: list[Paper],
        dimensions: Optional[list[str]] = None,
    ) -> dict:
        """
        跨多篇論文的維度比較分析。

        Args:
            papers: 要比較的論文（2-6 篇）
            dimensions: 比較維度（None 則讓 LLM 自動決定）

        Returns:
            {
                "dimensions": ["研究問題", "方法", ...],
                "papers": [{"title": ..., "authors": ..., "year": ...}, ...],
                "table": {
                    "研究問題": ["論文A的研究問題", "論文B的研究問題", ...],
                    "方法":     [...],
                },
                "synthesis": "綜合比較分析文字"
            }
        """
        if not self.client:
            raise ValueError("需要 Gemini API Key")
        if len(papers) < 2:
            raise ValueError("至少需要 2 篇論文才能比較")

        papers_ctx = "\n\n".join(
            f"【論文{i+1}】{p.title}（{p.year or '年份不明'}）\n摘要：{(p.abstract or '無摘要')[:600]}"
            for i, p in enumerate(papers)
        )

        if dimensions:
            dims_str = "、".join(dimensions)
            dim_instruction = f"請沿著以下維度進行比較：{dims_str}"
        else:
            dim_instruction = "請自動選擇 5-7 個最有意義的比較維度（如研究問題、方法、資料集、主要貢獻、局限性等）"

        prompt = f"""你是一位學術研究分析專家。請對以下 {len(papers)} 篇論文進行系統性比較分析。

{papers_ctx}

{dim_instruction}

請以 JSON 格式回傳比較結果：
{{
  "dimensions": ["維度1", "維度2", ...],
  "table": {{
    "維度1": ["論文1的內容（20字以內）", "論文2的內容", ...],
    "維度2": [...]
  }},
  "synthesis": "綜合分析：這些論文的共同點、主要差異、以及對領域的不同貢獻（150-200字）"
}}

只回傳 JSON。"""

        resp = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
        text = resp.text.strip()
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]

        data = json.loads(text.strip())

        return {
            "dimensions": data.get("dimensions", []),
            "papers": [
                {
                    "title": p.title,
                    "authors": "; ".join(p.authors[:2]) + (" et al." if len(p.authors) > 2 else "") if p.authors else "-",
                    "year": p.year,
                    "venue": p.venue,
                }
                for p in papers
            ],
            "table": data.get("table", {}),
            "synthesis": data.get("synthesis", ""),
        }
