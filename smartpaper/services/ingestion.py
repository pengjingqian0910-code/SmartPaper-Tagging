"""
XLSX / CSV 讀取模組
處理使用者上傳的論文標題清單，支援欄位自動偵測
"""

import csv
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from ..processing.cleaner import clean_paper_title


class XLSXIngestion:
    """XLSX / CSV 檔案讀取處理類"""

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"檔案不存在: {self.file_path}")

        self._is_csv = self.file_path.suffix.lower() == ".csv"

        if self._is_csv:
            self.workbook = None
            self.sheet = None
            self._csv_rows = self._load_csv()
        else:
            self.workbook = load_workbook(self.file_path, read_only=True)
            self.sheet = self.workbook.active
            self._csv_rows = None

    def _load_csv(self) -> list[list[str]]:
        with open(self.file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            return list(reader)

    def get_headers(self) -> list[tuple[int, str]]:
        """
        回傳所有欄位的 (索引, 名稱) 清單。
        CSV 用第一列做表頭；XLSX 用第一行的儲存格值。
        """
        if self._is_csv:
            if not self._csv_rows:
                return []
            return [(i, name.strip()) for i, name in enumerate(self._csv_rows[0]) if name.strip()]
        else:
            first_row = next(self.sheet.iter_rows(min_row=1, max_row=1))
            result = []
            for i, cell in enumerate(first_row):
                name = str(cell.value).strip() if cell.value else f"欄位{chr(65 + i)}"
                result.append((i, name))
            return result

    def detect_columns(self) -> dict[str, int]:
        """
        自動偵測欄位對應，回傳 {field: col_index}。
        支援 Zotero 匯出格式（Author / Publication Title / Publication Year / Abstract Note）。
        """
        keywords: dict[str, list[str]] = {
            "title": ["title", "論文標題", "標題", "paper title", "name", "paper"],
            "doi": ["doi"],
            "abstract": ["abstract note", "abstract", "摘要", "summary", "description"],
            "tags": ["tags", "標籤", "keywords", "關鍵詞", "keyword"],
            "authors": ["author", "authors", "作者"],
            "venue": ["publication title", "journal", "venue", "期刊", "會議", "conference"],
            "year": ["publication year", "year", "年份", "published"],
        }

        column_map: dict[str, int] = {}
        for idx, name in self.get_headers():
            name_lower = name.lower()
            for field, kws in keywords.items():
                if field not in column_map and any(kw in name_lower for kw in kws):
                    column_map[field] = idx
                    break

        return column_map

    def read_papers_by_index(
        self,
        title_col: int,
        abstract_col: int = -1,
        doi_col: int = -1,
        tags_col: int = -1,
        authors_col: int = -1,
        venue_col: int = -1,
        year_col: int = -1,
    ) -> list[dict]:
        """
        用欄位索引讀取論文資料，回傳 dict 清單。
        傳 -1 表示忽略該欄位。支援 Zotero 匯出格式。
        """
        papers = []

        if self._is_csv:
            data_rows: list = self._csv_rows[1:]

            def get_cell(row, idx):
                return row[idx] if 0 <= idx < len(row) and row[idx] else None
        else:
            data_rows = list(self.sheet.iter_rows(min_row=2, values_only=True))

            def get_cell(row, idx):
                return row[idx] if 0 <= idx < len(row) and row[idx] is not None else None

        for row in data_rows:
            title_val = get_cell(row, title_col)
            if not title_val:
                continue
            title = clean_paper_title(str(title_val))
            if not title:
                continue

            paper: dict = {"title": title}

            if abstract_col >= 0:
                val = get_cell(row, abstract_col)
                if val:
                    paper["abstract"] = str(val).strip()

            if doi_col >= 0:
                val = get_cell(row, doi_col)
                if val:
                    paper["doi"] = str(val).strip()

            if tags_col >= 0:
                val = get_cell(row, tags_col)
                if val:
                    paper["tags"] = [t.strip() for t in str(val).split(",") if t.strip()]

            if authors_col >= 0:
                val = get_cell(row, authors_col)
                if val:
                    # Zotero 格式："Last, First; Last2, First2"
                    parts = [a.strip() for a in str(val).split(";") if a.strip()]
                    paper["authors"] = parts

            if venue_col >= 0:
                val = get_cell(row, venue_col)
                if val:
                    paper["venue"] = str(val).strip()

            if year_col >= 0:
                val = get_cell(row, year_col)
                if val:
                    try:
                        paper["year"] = int(str(val).strip()[:4])
                    except (ValueError, TypeError):
                        pass

            papers.append(paper)

        return papers

    def read_titles(
        self,
        title_column: str = "A",
        start_row: int = 2,
        header_row: Optional[int] = 1,
    ) -> list[str]:
        """向後相容：以欄位字母讀取標題（供 CLI 使用）"""
        col_idx = ord(title_column.upper()) - ord("A")
        titles = []

        if self._is_csv:
            for row in self._csv_rows[start_row - 1:]:
                if col_idx < len(row) and row[col_idx]:
                    title = clean_paper_title(row[col_idx])
                    if title:
                        titles.append(title)
        else:
            for row in self.sheet.iter_rows(min_row=start_row):
                if col_idx < len(row):
                    cell_value = row[col_idx].value
                    if cell_value:
                        title = clean_paper_title(str(cell_value))
                        if title:
                            titles.append(title)

        return titles

    def get_row_count(self) -> int:
        if self._is_csv:
            return max(0, len(self._csv_rows) - 1)
        return self.sheet.max_row - 1 if self.sheet.max_row > 1 else 0

    def close(self) -> None:
        if self.workbook:
            self.workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
