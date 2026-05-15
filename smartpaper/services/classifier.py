"""
論文分類服務模組
根據用戶提供的主題關鍵字，將論文分類到相應的主題中

支援三種分類方式：
1. 語意搜尋（快速，直接搜摘要向量）
2. 兩階段 RAG（先搜標題，再用 LLM 分析摘要）
3. 純 LLM 分類（最精確但最慢）
"""

import json
import re
from typing import Optional, Callable, TYPE_CHECKING
from google import genai

from ..database.sqlite_db import SQLiteDB
from ..database.vector_db import VectorDB
from ..config import GEMINI_API_KEY, GEMINI_MODEL
from ..models import Paper
from .reranker import Reranker
from ..api.crossref import CrossrefAPI
from ..api.arxiv import ArxivAPI

if TYPE_CHECKING:
    from ..skills import SkillConfig


class ClassificationService:
    """論文分類服務"""

    def __init__(
        self,
        sqlite_db: Optional[SQLiteDB] = None,
        vector_db: Optional[VectorDB] = None,
        api_key: Optional[str] = None,
        skill: Optional["SkillConfig"] = None,
    ):
        """
        初始化分類服務

        Args:
            sqlite_db: SQLite 資料庫實例
            vector_db: 向量資料庫實例
            api_key: Gemini API Key
            skill: 專家角色設定
        """
        self.sqlite_db = sqlite_db or SQLiteDB()
        self.vector_db = vector_db or VectorDB()
        self.skill = skill
        self.reranker = Reranker()
        self.crossref = CrossrefAPI()
        self.arxiv = ArxivAPI()

        # 初始化 Gemini（用於 RAG 總結）
        self.api_key = api_key or GEMINI_API_KEY
        if self.api_key:
            self.client = genai.Client(api_key=self.api_key)
        else:
            self.client = None

    def _fetch_abstract_fallback(self, paper: Paper) -> str:
        """
        三段式補齊摘要：
        1. Crossref by DOI（最精確）
        2. Crossref by title（備援）
        3. Arxiv by title（開放取用論文）

        若成功找到摘要，同步回寫 DB 並回傳文字；否則回傳空字串。
        """
        # ── 1. Crossref by DOI
        crossref_result = None
        if paper.doi:
            try:
                crossref_result = self.crossref.get_by_doi(paper.doi)
            except Exception:
                pass

        # ── 2. Crossref by title
        if crossref_result is None or not crossref_result.abstract:
            try:
                crossref_result = self.crossref.search_by_title(paper.title)
            except Exception:
                pass

        if crossref_result and crossref_result.abstract:
            paper.abstract = crossref_result.abstract
            if not paper.venue and crossref_result.journal:
                paper.venue = crossref_result.journal
            if not paper.doi and crossref_result.doi:
                paper.doi = crossref_result.doi
            if paper.id:
                self.sqlite_db.update(paper)
            return paper.abstract

        # ── 3. Arxiv by title
        try:
            arxiv_result = self.arxiv.search_by_title(paper.title)
            if arxiv_result and arxiv_result.get("abstract"):
                paper.abstract = arxiv_result["abstract"]
                if paper.id:
                    self.sqlite_db.update(paper)
                return paper.abstract
        except Exception:
            pass

        return ""

    def _check_paper_relevance_title_only(self, topic: str, title: str) -> dict:
        """僅憑標題用 LLM 判斷相關性（無摘要時的備援，信心度上限 0.55）"""
        prompt = f"""根據論文標題，判斷這篇論文是否可能屬於「{topic}」主題。
注意：目前只有標題資訊，無法取得摘要，請保守估計。

論文標題：{title}

請以 JSON 格式回答（若不確定，傾向 false）：
{{
    "is_relevant": true或false,
    "confidence": 0.0到0.55之間的數字,
    "reason": "簡短說明（20字以內）",
    "topic_summary": ""
}}

只回傳 JSON，不要有其他文字。"""
        try:
            response = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            text = response.text.strip()
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]
            data = json.loads(text.strip())
            return {
                "is_relevant": bool(data.get("is_relevant", False)),
                "confidence": min(float(data.get("confidence", 0.3)), 0.55),
                "reason": data.get("reason", "（僅依標題判斷）"),
                "topic_summary": "",
            }
        except Exception as e:
            return {"is_relevant": False, "confidence": 0.0, "reason": f"分析失敗: {str(e)[:40]}", "topic_summary": ""}

    def classify_by_topics(
        self,
        topics: list[str],
        min_score: float = 0.3,
        max_papers_per_topic: int = 50,
    ) -> dict[str, list[dict]]:
        """
        根據主題關鍵字分類所有論文（使用語意搜尋）

        Args:
            topics: 主題關鍵字列表，例如 ["Machine Learning", "Healthcare", "NLP"]
            min_score: 最低相似度分數（0-1），低於此分數的不納入該分類
            max_papers_per_topic: 每個主題最多收錄的論文數量

        Returns:
            分類結果字典：
            {
                "Machine Learning": [{"paper": Paper, "score": 0.85}, ...],
                "Healthcare": [...],
                "_unclassified": [...]  # 未被分類的論文
            }
        """
        # 取得所有論文
        all_papers = self.sqlite_db.get_all(limit=1000)
        paper_dict = {p.id: p for p in all_papers}

        # 記錄每篇論文被分到哪些主題
        paper_classifications = {p.id: [] for p in all_papers}

        # 分類結果
        results = {topic: [] for topic in topics}
        results["_unclassified"] = []

        # 對每個主題進行語意搜尋
        for topic in topics:
            # 使用主題作為查詢，搜尋相關論文
            search_results = self.vector_db.search(
                query=topic,
                n_results=max_papers_per_topic,
            )

            for sr in search_results:
                paper_id = sr["paper_id"]
                score = sr["score"]

                if score >= min_score and paper_id in paper_dict:
                    results[topic].append({
                        "paper": paper_dict[paper_id],
                        "score": score,
                    })
                    paper_classifications[paper_id].append({
                        "topic": topic,
                        "score": score,
                    })

        # 找出未被分類的論文
        for paper_id, classifications in paper_classifications.items():
            if not classifications:
                results["_unclassified"].append({
                    "paper": paper_dict[paper_id],
                    "score": 0.0,
                })

        return results

    def classify_two_stage(
        self,
        topics: list[str],
        progress_callback: Optional[Callable[[str, int, int], None]] = None,
    ) -> dict[str, list[dict]]:
        """
        兩階段 RAG 分類（你要求的方式）：
        1. 先搜尋標題，找出候選論文
        2. 取得摘要，用 LLM 分析確認是否屬於該主題

        Args:
            topics: 主題關鍵字列表
            progress_callback: 進度回調函數 (topic, current, total)

        Returns:
            分類結果字典
        """
        if not self.client:
            raise ValueError("兩階段分類需要 Gemini API Key")

        all_papers = self.sqlite_db.get_all(limit=1000)
        paper_dict = {p.id: p for p in all_papers}

        # 記錄每篇論文被分到哪些主題
        paper_assigned = {p.id: False for p in all_papers}

        results = {topic: [] for topic in topics}
        results["_unclassified"] = []

        total_topics = len(topics)

        for topic_idx, topic in enumerate(topics):
            if progress_callback:
                progress_callback(topic, topic_idx + 1, total_topics)

            # ====== 階段 1：搜尋標題找候選論文 ======
            # 方法 A：關鍵字搜尋（標題包含關鍵字）
            keyword_matches = self.sqlite_db.search_by_title(topic)

            # 方法 B：也搜尋標籤
            tag_matches = self.sqlite_db.get_by_tag(topic)

            # 合併候選論文（去重）
            candidate_ids = set()
            candidates = []

            for paper in keyword_matches + tag_matches:
                if paper.id not in candidate_ids:
                    candidate_ids.add(paper.id)
                    candidates.append(paper)

            # 如果關鍵字搜尋結果太少，用語意搜尋補充
            if len(candidates) < 5:
                semantic_results = self.vector_db.search(query=topic, n_results=20)
                for sr in semantic_results:
                    paper_id = sr["paper_id"]
                    if paper_id not in candidate_ids and paper_id in paper_dict:
                        candidate_ids.add(paper_id)
                        candidates.append(paper_dict[paper_id])

            # ====== Re-ranking：對候選論文重新排序 ======
            if candidates:
                rerank_inputs = [
                    {
                        "paper": p,
                        "document": f"{p.title}. {p.abstract[:300] if p.abstract else ''}",
                    }
                    for p in candidates
                ]
                try:
                    reranked = self.reranker.rerank(
                        query=topic,
                        candidates=rerank_inputs,
                        text_key="document",
                        top_k=30,
                    )
                    candidates = [r["paper"] for r in reranked]
                except Exception as e:
                    print(f"Re-ranking 失敗，使用原始順序: {e}")

            # ====== 階段 2：用 LLM 分析摘要確認分類 ======
            for paper in candidates:
                # 若無摘要，先嘗試從 Crossref 補齊
                if not paper.abstract:
                    self._fetch_abstract_fallback(paper)

                if paper.abstract:
                    relevance = self._check_paper_relevance(
                        topic=topic,
                        title=paper.title,
                        abstract=paper.abstract,
                    )
                else:
                    # 實在找不到摘要，僅憑標題判斷
                    relevance = self._check_paper_relevance_title_only(topic, paper.title)

                if relevance["is_relevant"]:
                    results[topic].append({
                        "paper": paper,
                        "score": relevance["confidence"],
                        "reason": relevance["reason"],
                        "topic_summary": relevance["topic_summary"],
                    })
                    paper_assigned[paper.id] = True

        # 找出未被分類的論文
        for paper_id, assigned in paper_assigned.items():
            if not assigned:
                results["_unclassified"].append({
                    "paper": paper_dict[paper_id],
                    "score": 0.0,
                })

        return results

    def _check_paper_relevance(
        self,
        topic: str,
        title: str,
        abstract: str,
        include_summary: bool = True,
    ) -> dict:
        """
        使用 LLM 檢查論文是否與主題相關，並生成關聯摘要（RAG 的核心）

        Args:
            topic: 主題關鍵字
            title: 論文標題
            abstract: 論文摘要
            include_summary: 是否包含詳細的關聯摘要

        Returns:
            {
                "is_relevant": True/False,
                "confidence": 0.85,
                "reason": "簡短理由",
                "topic_summary": "這篇論文關於該主題的詳細說明..."
            }
        """
        # 摘要太長就截斷
        abstract_text = abstract[:1500] if abstract else "無摘要"

        # 使用 skill 的判斷標準（若有）
        if self.skill is not None:
            criteria = self.skill.classification_criteria
            role_prefix = self.skill.system_prompt + "\n\n"
        else:
            criteria = (
                "1. 論文的研究主題是否與該主題直接相關\n"
                "2. 論文使用的方法或技術是否屬於該領域\n"
                "3. 論文的應用場景是否涉及該主題"
            )
            role_prefix = ""

        prompt = f"""{role_prefix}請判斷以下論文是否屬於「{topic}」這個主題。

論文標題：{title}

論文摘要：
{abstract_text}

請分析這篇論文的內容，判斷它是否與「{topic}」主題相關。

判斷標準：
{criteria}

請以 JSON 格式回答：
{{
    "is_relevant": true或false,
    "confidence": 0.0到1.0之間的數字,
    "reason": "簡短說明判斷理由（30字以內）",
    "topic_summary": "詳細說明這篇論文與「{topic}」的關聯性。如果相關，請具體描述論文中哪些內容涉及{topic}，以及如何涉及（約100-150字）。如果不相關，此欄位留空。"
}}

只回傳 JSON，不要有其他文字。"""

        try:
            response = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            response_text = response.text.strip()

            # 清理 markdown 格式
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0]

            data = json.loads(response_text.strip())

            return {
                "is_relevant": bool(data.get("is_relevant", False)),
                "confidence": float(data.get("confidence", 0.5)),
                "reason": data.get("reason", ""),
                "topic_summary": data.get("topic_summary", "") if include_summary else "",
            }

        except Exception as e:
            print(f"LLM 檢查相關性失敗: {e}")
            return {
                "is_relevant": False,
                "confidence": 0.0,
                "reason": f"分析失敗: {str(e)[:50]}",
                "topic_summary": "",
            }

    def classify_with_llm(
        self,
        topics: list[str],
        batch_size: int = 10,
    ) -> dict[str, list[dict]]:
        """
        使用 LLM 進行多標籤分類（一篇論文可同時屬於多個主題）

        Args:
            topics: 主題關鍵字列表
            batch_size: 每次處理的論文數量

        Returns:
            分類結果字典（每個主題包含有符合的論文，論文可重複出現在多個主題）
        """
        if not self.client:
            raise ValueError("Gemini API Key 未設定，無法使用 LLM 分類")

        all_papers = self.sqlite_db.get_all(limit=1000)
        results = {topic: [] for topic in topics}
        results["_unclassified"] = []

        for i in range(0, len(all_papers), batch_size):
            batch = all_papers[i:i + batch_size]

            for paper in batch:
                if not paper.abstract:
                    self._fetch_abstract_fallback(paper)

                # 一次評估所有主題，回傳多標籤結果
                classifications = self._classify_single_paper_multilabel(paper, topics)

                if not classifications:
                    results["_unclassified"].append({"paper": paper, "score": 0.0})
                else:
                    for cls in classifications:
                        topic = cls["topic"]
                        if topic in results:
                            results[topic].append({
                                "paper": paper,
                                "score": cls["confidence"],
                            })

        return results

    def _classify_single_paper_multilabel(
        self,
        paper: Paper,
        topics: list[str],
        min_confidence: float = 0.3,
    ) -> list[dict]:
        """
        使用 LLM 對單篇論文進行多標籤分類

        Args:
            paper: 論文物件
            topics: 所有主題列表
            min_confidence: 最低信心門檻，低於此值視為不相關

        Returns:
            [{"topic": "ML", "confidence": 0.85}, {"topic": "Healthcare", "confidence": 0.42}, ...]
            相關程度低於 min_confidence 的主題不列入
        """
        topics_list = "\n".join(f"- {t}" for t in topics)

        prompt = f"""你是一位論文分類專家。請分析以下論文對各個主題的相關程度。

論文標題：{paper.title}
論文摘要：{paper.abstract[:800] if paper.abstract else "無摘要"}

需要評估的主題：
{topics_list}

評分標準：
- 0.8 以上：論文的核心研究主題，深度涉及
- 0.5～0.8：顯著相關，論文有實質討論
- 0.3～0.5：部分相關，論文有提及或輕度涉及
- 0.3 以下：不相關，無需列入

請對每個主題給出相關程度分數，只列出分數 ≥ {min_confidence} 的主題。

回傳 JSON（若所有主題都不相關，回傳空列表）：
{{"classifications": [{{"topic": "主題名稱", "confidence": 0.85}}, {{"topic": "另一主題", "confidence": 0.42}}]}}

只回傳 JSON，不要有其他文字。"""

        try:
            response = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            response_text = response.text.strip()

            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0]

            data = json.loads(response_text.strip())
            raw = data.get("classifications", [])

            # 過濾不合法主題並確保分數在範圍內
            valid_topics = set(topics)
            return [
                {"topic": c["topic"], "confidence": float(c["confidence"])}
                for c in raw
                if c.get("topic") in valid_topics
                and float(c.get("confidence", 0)) >= min_confidence
            ]
        except Exception as e:
            print(f"LLM 多標籤分類失敗: {e}")
            return []

    def build_paper_centric_view(
        self,
        classifications: dict[str, list[dict]],
        min_score: float = 0.0,
    ) -> list[dict]:
        """
        將 topic → papers 結構轉換成 paper → topics 的多標籤視圖

        Args:
            classifications: get_classification_report 產生的 topics 字典
            min_score: 只納入分數高於此門檻的主題

        Returns:
            [
                {
                    "paper": Paper,
                    "classifications": [
                        {"topic": "Machine Learning", "score": 0.68},
                        {"topic": "Healthcare", "score": 0.42},
                    ]
                },
                ...
            ]
            按最高分主題的分數遞減排序
        """
        paper_map: dict[int, dict] = {}

        for topic, papers in classifications.items():
            if topic == "_unclassified":
                continue
            for item in papers:
                paper = item["paper"]
                score = item["score"]
                if score < min_score:
                    continue

                pid = paper.id
                if pid not in paper_map:
                    paper_map[pid] = {"paper": paper, "classifications": []}

                paper_map[pid]["classifications"].append({
                    "topic": topic,
                    "score": score,
                })

        result = list(paper_map.values())
        for entry in result:
            entry["classifications"].sort(key=lambda x: x["score"], reverse=True)

        result.sort(
            key=lambda x: x["classifications"][0]["score"] if x["classifications"] else 0,
            reverse=True,
        )
        return result

    def generate_topic_summary(
        self,
        topic: str,
        papers: list[dict],
        max_papers: int = 10,
    ) -> str:
        """
        使用 RAG 為某個主題的論文生成總結

        Args:
            topic: 主題名稱
            papers: 該主題下的論文列表 [{"paper": Paper, "score": float}, ...]
            max_papers: 用於生成總結的最大論文數

        Returns:
            總結文字
        """
        if not self.client:
            return "（需要 Gemini API Key 才能生成總結）"

        if not papers:
            return f"「{topic}」主題下沒有論文。"

        # 取前 N 篇論文建構上下文
        top_papers = sorted(papers, key=lambda x: x["score"], reverse=True)[:max_papers]

        context_parts = []
        for i, item in enumerate(top_papers, 1):
            paper = item["paper"]
            abstract_preview = paper.abstract[:500] if paper.abstract else "無摘要"
            context_parts.append(
                f"{i}. 標題：{paper.title}\n   摘要：{abstract_preview}"
            )

        context = "\n\n".join(context_parts)

        summary_style = (
            self.skill.summary_style
            if self.skill is not None
            else "從研究目的、研究方法、主要發現三個維度進行總結"
        )

        prompt = f"""你是一位學術研究助理。以下是「{topic}」主題下的相關論文，請幫我總結這些論文的共同研究方向、主要貢獻和趨勢。

論文資料：
{context}

總結風格：{summary_style}

請用繁體中文回答，簡潔扼要（約 200-300 字）。"""

        try:
            response = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            return response.text.strip()
        except Exception as e:
            return f"生成總結時發生錯誤: {e}"

    def get_classification_report(
        self,
        topics: list[str],
        method: str = "semantic",
        include_summary: bool = True,
        progress_callback: Optional[Callable[[str, int, int], None]] = None,
    ) -> dict:
        """
        生成完整的分類報告

        Args:
            topics: 主題關鍵字列表
            method: 分類方法
                - "semantic": 語意搜尋（快速，直接搜摘要）
                - "two_stage": 兩階段 RAG（先搜標題，再用 LLM 分析摘要）★推薦
                - "llm": 純 LLM 分類（最精確但最慢）
            include_summary: 是否包含各主題的 RAG 總結
            progress_callback: 進度回調函數

        Returns:
            完整分類報告
        """
        # 執行分類
        if method == "two_stage":
            classifications = self.classify_two_stage(topics, progress_callback)
        elif method == "llm":
            classifications = self.classify_with_llm(topics)
        else:  # semantic
            classifications = self.classify_by_topics(topics)

        # 建構報告
        report = {
            "topics": {},
            "statistics": {
                "total_papers": 0,
                "classified_papers": 0,
                "unclassified_papers": 0,
            },
        }

        total_classified = 0

        for topic, papers in classifications.items():
            if topic == "_unclassified":
                report["statistics"]["unclassified_papers"] = len(papers)
                report["unclassified"] = [
                    {"title": p["paper"].title, "id": p["paper"].id}
                    for p in papers
                ]
                continue

            report["topics"][topic] = {
                "count": len(papers),
                "papers": [
                    {
                        "id": p["paper"].id,
                        "title": p["paper"].title,
                        "score": round(p["score"], 3),
                        "tags": p["paper"].tags,
                        "venue": p["paper"].venue,
                        "citation_count": p["paper"].citation_count,
                        "reason": p.get("reason", ""),
                        "topic_summary": p.get("topic_summary", ""),
                    }
                    for p in sorted(papers, key=lambda x: x["score"], reverse=True)
                ],
            }

            # 生成主題總結
            if include_summary and papers:
                report["topics"][topic]["summary"] = self.generate_topic_summary(
                    topic, papers
                )

            total_classified += len(papers)

        report["statistics"]["total_papers"] = self.sqlite_db.count()
        report["statistics"]["classified_papers"] = total_classified

        # 保留原始分類（含 Paper 物件），供 UI 呼叫 analyze_research_gaps 使用
        report["_raw_classifications"] = {
            topic: papers
            for topic, papers in classifications.items()
            if topic != "_unclassified"
        }

        # 論文視角：每篇論文列出所有符合的主題與分數
        paper_centric = self.build_paper_centric_view(classifications)
        report["paper_multi_label"] = [
            {
                "id": entry["paper"].id,
                "title": entry["paper"].title,
                "tags": entry["paper"].tags,
                "classifications": entry["classifications"],  # [{"topic": ..., "score": ...}]
            }
            for entry in paper_centric
        ]

        return report

    def analyze_research_gaps(
        self,
        topic: str,
        papers: list[dict],
        max_papers: int = 15,
    ) -> str:
        """
        分析某主題下的研究缺口，找出尚未被充分探索的方向

        Args:
            topic: 主題名稱
            papers: 論文列表 [{"paper": Paper, "score": float}, ...]
            max_papers: 用於分析的最大論文數（避免 prompt 過長）

        Returns:
            研究缺口分析文字（繁體中文）
        """
        if not self.client:
            return "（需要 Gemini API Key 才能執行缺口分析）"
        if not papers:
            return f"「{topic}」主題下沒有足夠論文可供分析。"

        top_papers = sorted(papers, key=lambda x: x["score"], reverse=True)[:max_papers]

        context_parts = []
        for i, item in enumerate(top_papers, 1):
            paper = item["paper"]
            abstract = (paper.abstract or "無摘要")[:400]
            venue = f"（{paper.venue}）" if paper.venue else ""
            context_parts.append(f"{i}. {paper.title}{venue}\n   {abstract}")

        context = "\n\n".join(context_parts)

        prompt = f"""你是一位資深學術研究顧問。以下是「{topic}」領域現有的相關論文：

{context}

請根據這些論文，進行研究缺口分析：

**1. 已被充分研究的方向**（2-3點，各一句話）
現有論文主要涵蓋哪些面向？

**2. 明顯的研究缺口**（3-5點，每點說明缺乏的原因與重要性）
哪些重要問題在現有論文中被忽略或不足？

**3. 建議研究方向**（3-5個具體方向，每個包含：研究問題 + 可能方法）
針對上述缺口，提出具體可執行的研究方向。

請用繁體中文回答，約 350-450 字。"""

        try:
            response = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            return response.text.strip()
        except Exception as e:
            return f"分析時發生錯誤：{e}"

    def suggest_topics(self, num_topics: int = 5) -> list[str]:
        """
        根據現有論文標籤，建議可能的分類主題

        Args:
            num_topics: 建議的主題數量

        Returns:
            建議的主題列表
        """
        # 取得所有標籤
        all_tags = self.sqlite_db.get_all_tags()

        if not all_tags:
            return []

        # 如果沒有 LLM，直接返回最常見的標籤
        if not self.client:
            return all_tags[:num_topics]

        # 使用 LLM 從標籤中歸納出主題
        tags_str = ", ".join(all_tags[:50])  # 最多取 50 個標籤

        prompt = f"""以下是一個論文資料庫中的標籤列表：
{tags_str}

請根據這些標籤，歸納出 {num_topics} 個主要的研究主題/分類。
這些主題應該：
1. 能涵蓋大部分的標籤
2. 彼此之間有區別
3. 是有意義的學術研究領域

請以 JSON 格式回答：
{{"topics": ["主題1", "主題2", ...]}}

只回傳 JSON。"""

        try:
            response = self.client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            response_text = response.text.strip()

            import json
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0]
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0]

            data = json.loads(response_text.strip())
            return data.get("topics", all_tags[:num_topics])
        except Exception:
            return all_tags[:num_topics]

    def get_papers_sorted_by_tags(
        self,
        sort_order: str = "tag_count",
        tag_filter: Optional[str] = None,
    ) -> list[dict]:
        """
        取得論文並根據標籤排序

        Args:
            sort_order: 排序方式
                - "tag_count": 按標籤數量排序（多到少）
                - "tag_alpha": 按第一個標籤字母順序排序
                - "tag_group": 按標籤分組
            tag_filter: 只顯示包含此標籤的論文

        Returns:
            排序後的論文列表
        """
        all_papers = self.sqlite_db.get_all(limit=1000)

        # 篩選
        if tag_filter:
            all_papers = [p for p in all_papers if tag_filter in (p.tags or [])]

        # 排序
        if sort_order == "tag_count":
            # 按標籤數量排序（多到少）
            all_papers.sort(key=lambda p: len(p.tags or []), reverse=True)
        elif sort_order == "tag_alpha":
            # 按第一個標籤字母順序排序
            all_papers.sort(key=lambda p: (p.tags or [""])[0].lower() if p.tags else "zzz")
        elif sort_order == "tag_group":
            # 按標籤分組（會在下面處理）
            pass

        # 建構結果
        results = []
        for paper in all_papers:
            results.append({
                "id": paper.id,
                "title": paper.title,
                "tags": paper.tags or [],
                "tag_count": len(paper.tags or []),
                "abstract": paper.abstract,
                "doi": paper.doi,
            })

        return results

    def get_papers_grouped_by_tag(self) -> dict[str, list[dict]]:
        """
        將論文按標籤分組

        Returns:
            {
                "Machine Learning": [paper1, paper2, ...],
                "Healthcare": [paper3, paper4, ...],
                "_no_tags": [paper5, ...]  # 沒有標籤的論文
            }
        """
        all_papers = self.sqlite_db.get_all(limit=1000)
        all_tags = self.sqlite_db.get_all_tags()

        # 初始化分組
        grouped = {tag: [] for tag in all_tags}
        grouped["_no_tags"] = []

        # 分組論文
        for paper in all_papers:
            if not paper.tags:
                grouped["_no_tags"].append({
                    "id": paper.id,
                    "title": paper.title,
                    "tags": [],
                })
            else:
                for tag in paper.tags:
                    if tag in grouped:
                        grouped[tag].append({
                            "id": paper.id,
                            "title": paper.title,
                            "tags": paper.tags,
                        })

        # 移除空的分組
        grouped = {k: v for k, v in grouped.items() if v}

        # 按論文數量排序分組
        sorted_grouped = dict(
            sorted(grouped.items(), key=lambda x: len(x[1]), reverse=True)
        )

        return sorted_grouped

    def export_classification_report(
        self,
        report: dict,
        output_path: str,
    ) -> None:
        """
        將分類報告匯出到 Excel

        Args:
            report: get_classification_report() 返回的報告
            output_path: 輸出檔案路徑
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("需要安裝 openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()

        # 建立總覽頁
        ws_overview = wb.active
        ws_overview.title = "總覽"

        # 標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # 寫入統計
        ws_overview["A1"] = "分類統計"
        ws_overview["A1"].font = Font(bold=True, size=14)

        ws_overview["A3"] = "論文總數"
        ws_overview["B3"] = report["statistics"]["total_papers"]
        ws_overview["A4"] = "已分類"
        ws_overview["B4"] = report["statistics"]["classified_papers"]
        ws_overview["A5"] = "未分類"
        ws_overview["B5"] = report["statistics"]["unclassified_papers"]

        # 為每個主題建立工作表
        for topic, data in report["topics"].items():
            # 清理工作表名稱（Excel 有限制）
            sheet_name = topic[:30].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=sheet_name)

            # 寫入標題列
            headers = ["信心度", "論文標題", "標籤", "分類理由", "主題關聯摘要"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # 寫入論文資料
            for row, paper in enumerate(data["papers"], 2):
                ws.cell(row=row, column=1, value=f"{paper['score']:.1%}")
                ws.cell(row=row, column=2, value=paper["title"])
                ws.cell(row=row, column=3, value=", ".join(paper.get("tags", [])))
                ws.cell(row=row, column=4, value=paper.get("reason", ""))
                ws.cell(row=row, column=5, value=paper.get("topic_summary", ""))

            # 調整欄寬
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 25
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["E"].width = 60

            # 寫入主題總結（如果有的話）
            if "summary" in data and data["summary"]:
                last_row = len(data["papers"]) + 3
                ws.cell(row=last_row, column=1, value="主題總結：")
                ws.cell(row=last_row, column=1).font = Font(bold=True)
                ws.cell(row=last_row + 1, column=1, value=data["summary"])
                ws.merge_cells(
                    start_row=last_row + 1,
                    start_column=1,
                    end_row=last_row + 1,
                    end_column=5
                )

        wb.save(output_path)
