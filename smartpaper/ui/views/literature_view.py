"""
文獻分析視圖
兩個子頁：
1. 文獻回顧表格 — 勾選欄位 → LLM 萃取 → Excel
2. 論文比較     — 選 2-6 篇 → LLM 比較維度表
"""

import flet as ft
from typing import Optional

from ...services.literature_analyzer import (
    LiteratureAnalyzer, FIXED_COLUMNS, LLM_COLUMNS,
)
from ...database.sqlite_db import SQLiteDB
from ...config import GEMINI_API_KEY


# ── 共用色彩 ──────────────────────────────────────────────────────────────
_C_BORDER   = "#E2E8F0"
_C_BG       = "#F8FAFC"
_C_TITLE    = "#1E293B"
_C_META     = "#64748B"
_PANEL_W    = 380   # 左側控制欄寬度


def _section_title(text: str) -> ft.Text:
    return ft.Text(text, size=12, weight=ft.FontWeight.W_600, color=_C_TITLE)


def _card(content: ft.Control, **kwargs) -> ft.Container:
    return ft.Container(
        content=content,
        padding=12,
        border=ft.border.all(1, _C_BORDER),
        border_radius=10,
        bgcolor="#FFFFFF",
        **kwargs,
    )


class LiteratureView:
    """文獻分析主視圖"""

    def __init__(self, page: ft.Page):
        self.page = page
        self.analyzer: Optional[LiteratureAnalyzer] = None
        self.db = SQLiteDB()

    def build(self) -> ft.Control:
        self.analyzer = LiteratureAnalyzer()

        self.status = ft.Text("", size=12, color=ft.colors.GREY_600)
        self.progress = ft.ProgressRing(visible=False, width=18, height=18, stroke_width=2)

        self.tabs = ft.Tabs(
            selected_index=0,
            animation_duration=200,
            tabs=[
                ft.Tab(text="文獻回顧表格",  icon="table_chart",     content=self._build_review_tab()),
                ft.Tab(text="論文比較分析",  icon="compare_arrows",  content=self._build_compare_tab()),
            ],
            expand=True,
        )

        return ft.Column(
            [
                ft.Row([
                    ft.Column([
                        ft.Text("文獻分析", size=24, weight=ft.FontWeight.BOLD, color=_C_TITLE),
                        ft.Text("文獻回顧表格 · 論文比較分析", size=12, color=_C_META),
                    ], spacing=2),
                    ft.Container(expand=True),
                    ft.Row([self.progress, self.status], spacing=8),
                ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Divider(height=1, color=_C_BORDER),
                self.tabs,
            ],
            spacing=8,
            expand=True,
        )

    # ════════════════════════════════════════════════════════════
    # Tab 1: 文獻回顧表格  ── 左右分欄
    # ════════════════════════════════════════════════════════════

    def _build_review_tab(self) -> ft.Control:
        # ── 欄位勾選 ──
        self._fixed_checks: dict[str, ft.Checkbox] = {
            col: ft.Checkbox(label=col,
                             value=(col in ["標題", "作者", "年份", "期刊/會議"]))
            for col in FIXED_COLUMNS
        }
        self._llm_checks: dict[str, ft.Checkbox] = {
            col: ft.Checkbox(label=col,
                             value=(col in ["研究方法", "主要貢獻", "研究限制"]))
            for col in LLM_COLUMNS
        }
        self.custom_cols_input = ft.TextField(
            label="自訂欄位（逗號分隔）",
            hint_text="例如：倫理考量, 開源狀態",
            border_radius=8,
        )

        # ── 論文勾選清單 ──
        all_papers = self.db.get_all(limit=2000)
        self._all_review_papers: dict[int, object] = {p.id: p for p in all_papers}
        self._review_selected: dict[int, ft.Checkbox] = {}
        review_checkboxes = []
        for p in all_papers[:500]:
            cb = ft.Checkbox(
                label=f"[{p.id}] {p.title[:55]}{'…' if len(p.title) > 55 else ''}"
                      f" ({p.year or '?'})",
                value=False,
                on_change=lambda _: (self._update_review_count(), self.page.update()),
            )
            self._review_selected[p.id] = cb
            review_checkboxes.append(cb)

        self.review_papers_list = ft.Column(
            controls=review_checkboxes,
            scroll=ft.ScrollMode.AUTO,
            spacing=2,
            expand=True,
        )
        self.review_search = ft.TextField(
            label="篩選論文", prefix_icon="search",
            border_radius=8,
            on_change=self._on_review_search,
            height=42,
            content_padding=ft.padding.symmetric(horizontal=10, vertical=0),
        )
        self._review_selected_count = ft.Text("已選 0 篇", size=11, color=ft.colors.BLUE_700)

        def _select_all(e):
            for pid, cb in self._review_selected.items():
                if cb.visible:
                    cb.value = True
            self._update_review_count()
            self.page.update()

        def _deselect_all(e):
            for cb in self._review_selected.values():
                cb.value = False
            self._update_review_count()
            self.page.update()

        # ── 結果區 ──
        self.review_results = ft.Column(scroll=ft.ScrollMode.AUTO, spacing=4, expand=True)
        self._review_rows = None
        self.review_export_btn = ft.ElevatedButton(
            "匯出 Excel", icon="download",
            on_click=self._on_review_export,
            visible=False,
            style=ft.ButtonStyle(bgcolor=ft.colors.GREEN_700, color=ft.colors.WHITE),
        )
        self.review_file_picker = ft.FilePicker()
        self.review_file_picker.on_result = self._on_review_save
        self.page.overlay.append(self.review_file_picker)

        gen_btn = ft.ElevatedButton(
            "生成表格", icon="table_chart",
            on_click=self._on_gen_review,
            style=ft.ButtonStyle(bgcolor=ft.colors.BLUE_700, color=ft.colors.WHITE),
        )

        # ── 左側控制欄 ──
        left_panel = ft.Container(
            content=ft.Column([
                # 欄位選擇
                _card(ft.Column([
                    _section_title("固定欄位（資料庫取得）"),
                    ft.Row(list(self._fixed_checks.values()), wrap=True, spacing=4),
                    ft.Divider(height=6),
                    _section_title("AI 萃取欄位（從摘要生成）"),
                    # 快選 preset pills
                    ft.Row([
                        ft.Text("快選：", size=10, color=_C_META),
                        *[self._preset_pill(name, cols)
                          for name, cols in [
                              ("方法論", ["研究方法", "主要貢獻"]),
                              ("評估",   ["資料集", "評估指標"]),
                              ("全選",   list(self._llm_checks.keys())),
                              ("清空",   []),
                          ]],
                    ], spacing=4, wrap=True),
                    ft.Row(list(self._llm_checks.values()), wrap=True, spacing=4),
                    ft.Divider(height=6),
                    self.custom_cols_input,
                ], spacing=8)),

                # 論文選擇
                _card(ft.Column([
                    ft.Row([
                        _section_title("選擇論文"),
                        ft.Container(expand=True),
                        self._review_selected_count,
                    ]),
                    self.review_search,
                    ft.Row([
                        ft.TextButton("全選", on_click=_select_all),
                        ft.TextButton("全不選", on_click=_deselect_all),
                    ], spacing=4),
                    ft.Container(
                        content=self.review_papers_list,
                        border=ft.border.all(1, _C_BORDER),
                        border_radius=6,
                        padding=6,
                        height=280,
                        clip_behavior=ft.ClipBehavior.HARD_EDGE,
                    ),
                ], spacing=8)),

                # 操作按鈕
                ft.Row([gen_btn, self.review_export_btn], spacing=10),
            ], spacing=12, scroll=ft.ScrollMode.AUTO),
            width=_PANEL_W,
            padding=ft.padding.only(top=12, right=12),
        )

        # ── 右側結果欄 ──
        right_panel = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon("table_view", color=ft.colors.BLUE_700, size=16),
                    _section_title("生成結果"),
                ], spacing=8),
                ft.Divider(height=4),
                ft.Container(content=self.review_results, expand=True),
            ], spacing=8, expand=True),
            expand=True,
            padding=ft.padding.only(top=12, left=4),
            border=ft.border.only(left=ft.border.BorderSide(1, _C_BORDER)),
        )

        return ft.Container(
            content=ft.Row([left_panel, right_panel],
                           spacing=0, expand=True,
                           vertical_alignment=ft.CrossAxisAlignment.START),
            expand=True,
        )

    def _preset_pill(self, label: str, cols: list[str]) -> ft.Container:
        def _apply(e, c=cols):
            for col, cb in self._llm_checks.items():
                cb.value = (col in c) if c else False
            self.page.update()
        return ft.Container(
            content=ft.Text(label, size=10, color="#4F46E5"),
            bgcolor="#EEF2FF",
            border=ft.border.all(1, "#C7D2FE"),
            border_radius=6,
            padding=ft.padding.symmetric(horizontal=8, vertical=3),
            on_click=_apply,
            ink=True,
        )

    def _on_review_search(self, e):
        keyword = (e.control.value or "").lower()
        for pid, cb in self._review_selected.items():
            paper = self._all_review_papers.get(pid)
            if paper:
                cb.visible = not keyword or keyword in paper.title.lower()
        self.page.update()

    def _update_review_count(self):
        n = sum(1 for cb in self._review_selected.values() if cb.value)
        self._review_selected_count.value = f"已選 {n} 篇"

    def _on_gen_review(self, e):
        fixed  = [col for col, cb in self._fixed_checks.items() if cb.value]
        llm    = [col for col, cb in self._llm_checks.items() if cb.value]
        custom = [c.strip() for c in self.custom_cols_input.value.split(",") if c.strip()]
        if not fixed and not llm and not custom:
            self.status.value = "請至少勾選一個欄位"
            self.page.update()
            return

        papers = [
            self._all_review_papers[pid]
            for pid, cb in self._review_selected.items()
            if cb.value and pid in self._all_review_papers
        ]
        if not papers:
            self.status.value = "請勾選至少一篇論文"
            self.page.update()
            return

        self._set_busy(f"生成文獻回顧表格（共 {len(papers)} 篇）...")
        self.review_results.controls.clear()
        self.review_export_btn.visible = False
        self.page.update()

        def _progress(cur, total):
            self.status.value = f"分析中 [{cur}/{total}]..."
            self.page.update()

        import threading
        def _run():
            try:
                rows = self.analyzer.generate_review_table(
                    papers=papers,
                    fixed_cols=fixed,
                    llm_cols=llm,
                    custom_cols=custom if custom else None,
                    progress_callback=_progress if llm or custom else None,
                )
                self._review_rows = rows
                self._render_review_table(rows)
                self.review_export_btn.visible = True
                self.status.value = (
                    f"表格生成完成：{len(rows)} 篇 × "
                    f"{len(rows[0]) if rows else 0} 欄"
                )
            except Exception as ex:
                self.status.value = f"錯誤：{ex}"
            finally:
                self._set_idle()
        threading.Thread(target=_run, daemon=True).start()

    def _render_review_table(self, rows: list[dict]):
        if not rows:
            return
        headers = list(rows[0].keys())
        col_defs = [
            ft.DataColumn(ft.Text(h, size=11, weight=ft.FontWeight.BOLD))
            for h in headers
        ]
        data_rows = []
        for row in rows:
            cells = []
            for h in headers:
                val = str(row.get(h, "-"))
                cells.append(ft.DataCell(
                    ft.Container(
                        content=ft.Text(
                            val[:100] + ("…" if len(val) > 100 else ""),
                            size=10, tooltip=val,
                        ),
                        width=140,
                    )
                ))
            data_rows.append(ft.DataRow(cells=cells))

        table = ft.DataTable(
            columns=col_defs,
            rows=data_rows,
            border=ft.border.all(1, _C_BORDER),
            border_radius=6,
            heading_row_height=36,
            data_row_min_height=40,
            column_spacing=8,
        )
        self.review_results.controls.append(
            ft.Container(
                content=ft.Row([table], scroll=ft.ScrollMode.AUTO),
                expand=True,
            )
        )
        self.page.update()

    def _on_review_export(self, e):
        self.review_file_picker.save_file(
            allowed_extensions=["xlsx"],
            file_name="literature_review.xlsx",
            dialog_title="儲存文獻回顧表格",
        )

    def _on_review_save(self, e):
        if not e.path or not self._review_rows:
            return
        self._set_busy("匯出 Excel...")
        try:
            self.analyzer.save_review_table_xlsx(self._review_rows, e.path)
            self.status.value = f"已儲存：{e.path}"
        except Exception as ex:
            self.status.value = f"匯出失敗：{ex}"
        finally:
            self._set_idle()

    # ════════════════════════════════════════════════════════════
    # Tab 3: 論文比較分析  ── 左右分欄
    # ════════════════════════════════════════════════════════════

    def _build_compare_tab(self) -> ft.Control:
        all_papers = self.db.get_all(limit=1000)
        self._all_papers = {p.id: p for p in all_papers}
        self._compare_selected: dict[int, ft.Checkbox] = {}
        self._compare_selected_count = ft.Text("已選 0 篇", size=11, color=ft.colors.PURPLE_700)

        paper_checkboxes = []
        for p in all_papers[:200]:
            cb = ft.Checkbox(
                label=f"[{p.id}] {p.title[:50]}{'…' if len(p.title) > 50 else ''}"
                      f" ({p.year or '?'})",
                value=False,
                on_change=lambda _: (self._update_compare_count(), self.page.update()),
            )
            self._compare_selected[p.id] = cb
            paper_checkboxes.append(cb)

        self.compare_papers_list = ft.Column(
            controls=paper_checkboxes,
            scroll=ft.ScrollMode.AUTO,
            spacing=2,
            expand=True,
        )
        self.compare_dims_input = ft.TextField(
            label="自訂比較維度（逗號分隔，留空則 AI 自動決定）",
            hint_text="例如：研究問題, 方法, 資料集, 主要貢獻, 局限性",
            border_radius=8,
        )
        self.compare_search = ft.TextField(
            label="篩選論文", prefix_icon="search",
            border_radius=8, height=42,
            content_padding=ft.padding.symmetric(horizontal=10, vertical=0),
            on_change=self._on_compare_search,
        )

        self.compare_results = ft.Column(scroll=ft.ScrollMode.AUTO, spacing=6, expand=True)
        self._compare_data = None

        self.compare_export_btn = ft.OutlinedButton(
            "匯出比較表 Excel", icon="download",
            on_click=self._on_compare_export, visible=False,
        )
        self.compare_file_picker = ft.FilePicker()
        self.compare_file_picker.on_result = self._on_compare_save
        self.page.overlay.append(self.compare_file_picker)

        compare_btn = ft.ElevatedButton(
            "開始比較", icon="compare_arrows",
            on_click=self._on_compare,
            style=ft.ButtonStyle(bgcolor=ft.colors.PURPLE_700, color=ft.colors.WHITE),
        )

        # ── 左側控制欄 ──
        left_panel = ft.Container(
            content=ft.Column([
                _card(ft.Column([
                    ft.Row([
                        _section_title("選擇 2–6 篇論文"),
                        ft.Container(expand=True),
                        self._compare_selected_count,
                    ]),
                    self.compare_search,
                    ft.Container(
                        content=self.compare_papers_list,
                        border=ft.border.all(1, _C_BORDER),
                        border_radius=6,
                        padding=6,
                        height=340,
                        clip_behavior=ft.ClipBehavior.HARD_EDGE,
                    ),
                ], spacing=8)),

                _card(ft.Column([
                    _section_title("比較維度設定"),
                    self.compare_dims_input,
                ], spacing=8)),

                ft.Row([compare_btn, self.compare_export_btn], spacing=10),
            ], spacing=12, scroll=ft.ScrollMode.AUTO),
            width=_PANEL_W,
            padding=ft.padding.only(top=12, right=12),
        )

        # ── 右側結果欄 ──
        right_panel = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon("compare_arrows", color=ft.colors.PURPLE_700, size=16),
                    _section_title("比較結果"),
                ], spacing=8),
                ft.Divider(height=4),
                ft.Container(content=self.compare_results, expand=True),
            ], spacing=8, expand=True),
            expand=True,
            padding=ft.padding.only(top=12, left=4),
            border=ft.border.only(left=ft.border.BorderSide(1, _C_BORDER)),
        )

        return ft.Container(
            content=ft.Row([left_panel, right_panel],
                           spacing=0, expand=True,
                           vertical_alignment=ft.CrossAxisAlignment.START),
            expand=True,
        )

    def _update_compare_count(self):
        n = sum(1 for cb in self._compare_selected.values() if cb.value)
        self._compare_selected_count.value = f"已選 {n} 篇"

    def _on_compare_search(self, e):
        keyword = (e.control.value or "").lower()
        for pid, cb in self._compare_selected.items():
            paper = self._all_papers.get(pid)
            if paper:
                cb.visible = not keyword or keyword in paper.title.lower()
        self.page.update()

    def _on_compare(self, e):
        selected = [
            self._all_papers[pid]
            for pid, cb in self._compare_selected.items()
            if cb.value
        ]
        if len(selected) < 2:
            self.status.value = "請至少勾選 2 篇論文"
            self.page.update()
            return
        if len(selected) > 6:
            self.status.value = "最多選 6 篇論文（避免 prompt 過長）"
            self.page.update()
            return

        dims_raw = self.compare_dims_input.value.strip()
        dims = [d.strip() for d in dims_raw.split(",") if d.strip()] if dims_raw else None

        self._set_busy(f"比較 {len(selected)} 篇論文中...")
        self.compare_results.controls.clear()
        self.compare_export_btn.visible = False
        self.page.update()

        import threading
        def _run():
            try:
                result = self.analyzer.compare_papers(papers=selected, dimensions=dims)
                self._compare_data = (result, selected)
                self._render_comparison(result)
                self.compare_export_btn.visible = True
                self.status.value = f"比較完成：{len(result['dimensions'])} 個維度"
            except Exception as ex:
                self.status.value = f"錯誤：{ex}"
            finally:
                self._set_idle()
        threading.Thread(target=_run, daemon=True).start()

    def _render_comparison(self, result: dict):
        dims        = result.get("dimensions", [])
        papers_meta = result.get("papers", [])
        table_data  = result.get("table", {})
        synthesis   = result.get("synthesis", "")

        if not dims or not papers_meta:
            return

        # 論文索引
        index_items = [
            ft.Text(
                f"[{i+1}] {m['title']} ({m.get('year', '?')})"
                + (f" | {m['authors']}" if m.get('authors') and m['authors'] != '-' else ""),
                size=10, color=ft.colors.GREY_700,
            )
            for i, m in enumerate(papers_meta)
        ]

        # 比較表格
        headers = ["比較維度"] + [
            f"[{i+1}] {m['title'][:25]}…" for i, m in enumerate(papers_meta)
        ]
        col_defs = [
            ft.DataColumn(ft.Text(h, size=10, weight=ft.FontWeight.BOLD))
            for h in headers
        ]
        rows = []
        for dim in dims:
            values = table_data.get(dim, [])
            cells = [ft.DataCell(
                ft.Text(dim, size=10, weight=ft.FontWeight.W_600,
                        color=ft.colors.PURPLE_800)
            )]
            for j in range(len(papers_meta)):
                val = values[j] if j < len(values) else "-"
                cells.append(ft.DataCell(
                    ft.Container(
                        content=ft.Text(str(val), size=10, tooltip=str(val)),
                        width=150,
                    )
                ))
            rows.append(ft.DataRow(cells=cells))

        table = ft.DataTable(
            columns=col_defs,
            rows=rows,
            border=ft.border.all(1, _C_BORDER),
            border_radius=6,
            heading_row_height=40,
            data_row_min_height=44,
            column_spacing=8,
        )

        synthesis_card = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon("analytics", size=15, color=ft.colors.PURPLE),
                    ft.Text("綜合比較分析", size=12, weight=ft.FontWeight.W_600),
                ], spacing=6),
                ft.Text(synthesis, size=11, color=ft.colors.GREY_800, selectable=True),
            ], spacing=6),
            bgcolor=ft.colors.PURPLE_50,
            padding=12,
            border_radius=8,
        ) if synthesis else ft.Container()

        self.compare_results.controls.extend([
            _card(ft.Column(index_items, spacing=2)),
            ft.Container(height=4),
            # 橫向捲動的表格
            ft.Container(
                content=ft.Row([table], scroll=ft.ScrollMode.AUTO),
                border=ft.border.all(1, _C_BORDER),
                border_radius=8,
                padding=8,
            ),
            synthesis_card,
        ])
        self.page.update()

    def _on_compare_export(self, e):
        self.compare_file_picker.save_file(
            allowed_extensions=["xlsx"],
            file_name="paper_comparison.xlsx",
            dialog_title="儲存比較分析",
        )

    def _on_compare_save(self, e):
        if not e.path or not self._compare_data:
            return
        self._set_busy("匯出 Excel...")
        result, papers = self._compare_data
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "論文比較"

            dims        = result["dimensions"]
            papers_meta = result["papers"]
            table       = result["table"]

            header_fill = PatternFill(start_color="6B3FA0", end_color="6B3FA0",
                                      fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            headers = ["比較維度"] + [
                f"[{i+1}] {m['title'][:40]}" for i, m in enumerate(papers_meta)
            ]
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font   = header_font
                cell.fill   = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="center")

            for ri, dim in enumerate(dims, 2):
                ws.cell(row=ri, column=1, value=dim).font = Font(bold=True)
                for ci, val in enumerate(table.get(dim, []), 2):
                    ws.cell(row=ri, column=ci, value=val).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )

            if result.get("synthesis"):
                last = len(dims) + 3
                ws.cell(row=last, column=1, value="綜合分析").font = Font(bold=True)
                ws.merge_cells(start_row=last, start_column=2,
                               end_row=last, end_column=len(headers))
                ws.cell(row=last, column=2, value=result["synthesis"])

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 30
            wb.save(e.path)
            self.status.value = f"已儲存：{e.path}"
        except Exception as ex:
            self.status.value = f"匯出失敗：{ex}"
        finally:
            self._set_idle()

    # ── Helpers ───────────────────────────────────────────────────────────

    def _set_busy(self, msg: str):
        self.progress.visible = True
        self.status.value = msg
        self.page.update()

    def _set_idle(self):
        self.progress.visible = False
        self.page.update()
